#!/usr/bin/env python3
"""
Genera un file Excel filtrabile con la lista completa dei giocatori.

Colonne (in quest'ordine):
  1. Calciatore   (le "scommesse" senza storico sono prefissate con 🎲)
  2. Ruolo/i      (ruoli Mantra, es. "Dd/E")
  3. Squadra
  4. Valutazione  (1-20)
  5. Situazione infortuni

VALUTAZIONE 1-20 — riusa la logica di scripts/ranking_reparto.py:
  - i giocatori sono raggruppati nei 5 reparti (portieri, difensori,
    centrocampisti, esterni, trequartisti/attaccanti) per RUOLO PRIMARIO;
  - per i giocatori con storico si calcola lo stesso score di ranking_reparto
    (metriche + pesi del reparto + normalizzazione min-max dei componenti),
    normalizzato POI su 1-20 DENTRO al gruppo (min-max) -> l'ordine e' identico
    a quello prodotto da ranking_reparto per quel reparto;
  - le "scommesse" (senza storico) ricevono una stima prudente (banda 4-11)
    ricavata dall'FVM del gruppo, e il simbolo 🎲.

Il file ha autofilter sull'intestazione e la riga di intestazione congelata.

Uso:
    python3 scripts/genera_lista.py [--out cartella] [--data YYYY-MM-DD]
Output: <out>/lista_giocatori_<data>.xlsx  (stampa il percorso)
"""
import argparse
import csv
import datetime
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "scripts"))
import ranking_reparto as rr  # noqa: E402

DATASET = REPO / "data" / "dataset_unificato.csv"

# ruolo Mantra primario -> uno dei 5 gruppi (per la normalizzazione)
RUOLO_GRUPPO = {
    "por": "portieri",
    "dc": "difensori", "dd": "difensori", "ds": "difensori", "b": "difensori",
    "m": "centrocampisti", "c": "centrocampisti",
    "e": "esterni", "w": "esterni",
    "t": "attaccanti", "a": "attaccanti", "pc": "attaccanti",
}
# gruppo -> preset di ranking_reparto da cui prendere pesi e metrica di qualita'
GRUPPO_PRESET = {
    "portieri": "por", "difensori": "dif", "centrocampisti": "cen",
    "esterni": "att", "attaccanti": "att",
}
ORDINE = ["portieri", "difensori", "centrocampisti", "esterni", "attaccanti"]
ETICHETTA = {"portieri": "Portieri", "difensori": "Difensori",
             "centrocampisti": "Centrocampisti", "esterni": "Esterni",
             "attaccanti": "Trequartisti/Attaccanti"}


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _gruppo(ruolo_mantra: str) -> str:
    primo = (ruolo_mantra or "").split("/")[0].strip().lower()
    return RUOLO_GRUPPO.get(primo, "attaccanti")


def _val_1_20(valore, lo, hi):
    if hi is None or lo is None or hi == lo:
        return 13
    return round(1 + 19 * (valore - lo) / (hi - lo))


def _valuta_gruppo(giocatori: list[dict], preset: str):
    """Assegna la valutazione 1-20 ai giocatori di un gruppo.

    Riproduce lo scoring di ranking_reparto.classifica (stessi metriche, pesi e
    normalizzazione dei componenti), poi mappa lo score su 1-20 dentro al gruppo.
    Ritorna dict player_id -> (valutazione:int, scommessa:bool).
    """
    qfield = rr.REPARTI[preset]["qualita"]
    w_cont, w_qual, w_bonus, w_card, w_forza = rr.REPARTI[preset]["w"]

    proven = [g for g in giocatori if g["nessuno_storico"] != "1"]
    scommesse = [g for g in giocatori if g["nessuno_storico"] == "1"]
    out = {}

    if proven:
        met = [rr.metriche(g) for g in proven]
        n_cont = rr._norm([m["cont"] for m in met])
        n_qual = rr._norm([m[qfield] for m in met])
        n_bonus = rr._norm([m["bonus"] for m in met])
        n_card = rr._norm([m["card"] for m in met])
        n_forza = rr._norm([m["forza"] for m in met])
        scores = []
        for g, c, f, b, cm, fz in zip(proven, n_cont, n_qual, n_bonus, n_card, n_forza):
            s = w_cont * c + w_qual * f + w_bonus * b - w_card * cm + w_forza * fz
            s += {"crescente": 0.03, "calante": -0.03}.get(g["trend_fm"], 0.0)
            scores.append(s)
        lo, hi = min(scores), max(scores)
        for g, s in zip(proven, scores):
            out[g["player_id"]] = (_val_1_20(s, lo, hi), False)

    if scommesse:
        fvm = [_f(g["mantra_fvm"]) for g in scommesse]
        pres = [x for x in fvm if x is not None]
        lo, hi = (min(pres), max(pres)) if pres else (None, None)
        for g, x in zip(scommesse, fvm):
            if x is None or lo is None or hi == lo:
                v = 6
            else:
                v = round(4 + 7 * (x - lo) / (hi - lo))  # banda prudente 4-11
            out[g["player_id"]] = (v, True)
    return out


def costruisci_righe() -> list[dict]:
    with open(DATASET, encoding="utf-8") as f:
        players = list(csv.DictReader(f))

    # raggruppa e valuta per reparto
    per_gruppo = {g: [] for g in ORDINE}
    for p in players:
        per_gruppo[_gruppo(p["ruolo_mantra"])].append(p)
    valut = {}
    for g in ORDINE:
        valut.update(_valuta_gruppo(per_gruppo[g], GRUPPO_PRESET[g]))

    righe = []
    for g in ORDINE:
        for p in per_gruppo[g]:
            v, scommessa = valut.get(p["player_id"], (6, True))
            nome = ("🎲 " if scommessa else "") + p["nome"]
            ruoli = "/".join(r.upper() for r in p["ruolo_mantra"].split("/"))
            inf = ""
            if p.get("stato_infortunio"):
                inf = "⚠️ " + p.get("dettaglio_infortunio", "")
                if p.get("infortunio_aggiornato_il"):
                    inf += f"  (agg. {p['infortunio_aggiornato_il']})"
            righe.append({"_gruppo": g, "Calciatore": nome, "Ruolo/i": ruoli,
                          "Squadra": p["squadra"], "Valutazione": v,
                          "Situazione infortuni": inf})
    # ordina: per reparto (ordine fisso), poi valutazione desc, poi nome
    righe.sort(key=lambda r: (ORDINE.index(r["_gruppo"]), -r["Valutazione"], r["Calciatore"]))
    return righe


def scrivi_xlsx(righe: list[dict], out_path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    colonne = ["Calciatore", "Ruolo/i", "Squadra", "Valutazione", "Situazione infortuni"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Giocatori"

    # intestazione
    ws.append(colonne)
    intest_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = intest_fill
        cell.alignment = Alignment(vertical="center")

    for r in righe:
        ws.append([r[c] for c in colonne])

    # larghezze + wrap sulla colonna infortuni
    larghezze = {"A": 22, "B": 12, "C": 9, "D": 12, "E": 90}
    for col, w in larghezze.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        row[0].alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"                         # intestazione congelata
    ws.auto_filter.ref = f"A1:E{ws.max_row}"       # filtro automatico

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def genera(out_dir: Path | None = None, data: str | None = None) -> Path:
    data = data or datetime.date.today().isoformat()
    out_dir = out_dir or REPO
    righe = costruisci_righe()
    return scrivi_xlsx(righe, out_dir / f"lista_giocatori_{data}.xlsx")


def main():
    ap = argparse.ArgumentParser(description="Genera lista giocatori in Excel")
    ap.add_argument("--out", help="cartella di output (default: radice repo)")
    ap.add_argument("--data", help="data nel nome file (default: oggi)")
    args = ap.parse_args()
    path = genera(Path(args.out) if args.out else None, args.data)
    print(path)


if __name__ == "__main__":
    main()
